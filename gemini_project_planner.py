# Complete Agentic AI Project Planning System
# Requirements: pip install langchain google-generativeai pandas openpyxl streamlit python-docx PyPDF2

import streamlit as st
import pandas as pd
import json
import re
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import google.generativeai as genai
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.chains import LLMChain
from langchain.prompts import PromptTemplate
from langchain.agents import Tool, AgentExecutor, create_react_agent
from langchain.memory import ConversationBufferMemory
from langchain.schema import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import PyPDF2
from docx import Document as DocxDocument
from dotenv import load_dotenv
load_dotenv()

# =============================================================================
# AGENT CLASSES
# =============================================================================

class ProjectParserAgent:
    """Agent responsible for parsing and understanding project requirements"""
    
    def __init__(self, llm):
        self.llm = llm
        self.prompt = PromptTemplate(
            input_variables=["requirements"],
            template="""
            You are a project analysis expert. Parse the following project requirements and extract:
            
            1. Project Name
            2. Project Type (web app, mobile app, desktop, API, etc.)
            3. Key Features (list)
            4. Technical Requirements
            5. Estimated Complexity (Simple/Medium/Complex)
            6. Target Timeline
            7. Team Roles Needed
            8. Critical Constraints
            
            Requirements: {requirements}
            
            Return the analysis in JSON format:
            {{
                "project_name": "...",
                "project_type": "...",
                "key_features": [...],
                "technical_requirements": [...],
                "complexity": "...",
                "estimated_timeline": "...",
                "team_roles": [...],
                "constraints": [...]
            }}
            """
        )
        self.chain = LLMChain(llm=self.llm, prompt=self.prompt)
    
    def parse_requirements(self, requirements: str) -> Dict[str, Any]:
        """Parse project requirements and return structured data"""
        try:
            result = self.chain.run(requirements=requirements)
            # Clean and parse JSON
            json_match = re.search(r'\{.*\}', result, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
            else:
                return self._fallback_parsing(requirements)
        except Exception as e:
            st.error(f"Error parsing requirements: {str(e)}")
            return self._fallback_parsing(requirements)
    
    def _fallback_parsing(self, requirements: str) -> Dict[str, Any]:
        """Fallback parsing when JSON parsing fails"""
        return {
            "project_name": "Custom Project",
            "project_type": "Web Application",
            "key_features": ["User Interface", "Backend API", "Database"],
            "technical_requirements": ["Frontend Framework", "Backend Server", "Database"],
            "complexity": "Medium",
            "estimated_timeline": "4-6 weeks",
            "team_roles": ["Frontend Developer", "Backend Developer", "Designer"],
            "constraints": ["Budget", "Timeline"]
        }

class ProjectPlannerAgent:
    """Agent responsible for breaking down projects into tasks and creating schedules"""
    
    def __init__(self, llm):
        self.llm = llm
        self.prompt = PromptTemplate(
            input_variables=["project_data", "team_size"],
            template="""
            You are a project planning expert. Create a detailed project plan with tasks, dependencies, and timeline.
            
            Project Data: {project_data}
            Team Size: {team_size}
            
            Create tasks with the following structure for each task:
            - Task Name
            - Description
            - Duration in days
            - Dependencies (list of task IDs that must complete first)
            - Assigned Role
            - Priority (High/Medium/Low)
            - Category (Setup/Design/Backend/Frontend/Testing/Deployment)
            
            Return as JSON array:
            [
                {{
                    "id": 1,
                    "task_name": "...",
                    "description": "...",
                    "duration": 3,
                    "dependencies": [],
                    "assigned_to": "...",
                    "priority": "High",
                    "category": "Setup"
                }}
            ]
            
            Consider:
            - Logical task dependencies
            - Parallel work opportunities
            - Critical path optimization
            - Resource allocation
            """
        )
        self.chain = LLMChain(llm=self.llm, prompt=self.prompt)
    
    def create_project_plan(self, project_data: Dict[str, Any], team_size: int = 3) -> List[Dict[str, Any]]:
        """Create detailed project plan with tasks and dependencies"""
        try:
            result = self.chain.run(
                project_data=json.dumps(project_data, indent=2),
                team_size=team_size
            )
            
            # Extract JSON array
            json_match = re.search(r'\[.*\]', result, re.DOTALL)
            if json_match:
                tasks = json.loads(json_match.group())
                return self._add_dates_to_tasks(tasks)
            else:
                return self._create_fallback_plan(project_data)
                
        except Exception as e:
            st.error(f"Error creating project plan: {str(e)}")
            return self._create_fallback_plan(project_data)
    
    def _add_dates_to_tasks(self, tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Add start and end dates to tasks based on dependencies"""
        start_date = datetime.now().date()
        task_dates = {}
        
        for task in tasks:
            task_id = task['id']
            duration = task['duration']
            dependencies = task.get('dependencies', [])
            
            if not dependencies:
                # No dependencies, start immediately
                task_start = start_date
            else:
                # Start after all dependencies complete
                latest_end = start_date
                for dep_id in dependencies:
                    if dep_id in task_dates:
                        dep_end = task_dates[dep_id]['end_date']
                        if dep_end > latest_end:
                            latest_end = dep_end
                task_start = latest_end + timedelta(days=1)
            
            task_end = task_start + timedelta(days=duration - 1)
            
            task['start_date'] = task_start.strftime('%Y-%m-%d')
            task['end_date'] = task_end.strftime('%Y-%m-%d')
            
            task_dates[task_id] = {
                'start_date': task_start,
                'end_date': task_end
            }
        
        return tasks
    
    def _create_fallback_plan(self, project_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Create a basic fallback plan when AI generation fails"""
        base_tasks = [
            {
                "id": 1,
                "task_name": "Project Setup & Planning",
                "description": "Initialize project structure and define architecture",
                "duration": 2,
                "dependencies": [],
                "assigned_to": "Tech Lead",
                "priority": "High",
                "category": "Setup"
            },
            {
                "id": 2,
                "task_name": "Design & Prototyping",
                "description": "Create UI/UX designs and prototypes",
                "duration": 5,
                "dependencies": [1],
                "assigned_to": "Designer",
                "priority": "High",
                "category": "Design"
            },
            {
                "id": 3,
                "task_name": "Backend Development",
                "description": "Implement server-side logic and APIs",
                "duration": 8,
                "dependencies": [1],
                "assigned_to": "Backend Developer",
                "priority": "High",
                "category": "Backend"
            },
            {
                "id": 4,
                "task_name": "Frontend Development",
                "description": "Implement user interface and client-side logic",
                "duration": 6,
                "dependencies": [2],
                "assigned_to": "Frontend Developer",
                "priority": "Medium",
                "category": "Frontend"
            },
            {
                "id": 5,
                "task_name": "Integration & Testing",
                "description": "Integrate components and perform testing",
                "duration": 4,
                "dependencies": [3, 4],
                "assigned_to": "QA Engineer",
                "priority": "High",
                "category": "Testing"
            }
        ]
        
        return self._add_dates_to_tasks(base_tasks)

class ProjectOptimizerAgent:
    """Agent responsible for optimizing and refining project plans based on feedback"""
    
    def __init__(self, llm):
        self.llm = llm
        self.prompt = PromptTemplate(
            input_variables=["current_plan", "feedback", "constraints"],
            template="""
            You are a project optimization expert. Refine the current project plan based on user feedback and constraints.
            
            Current Plan: {current_plan}
            User Feedback: {feedback}
            Additional Constraints: {constraints}
            
            Adjust the plan considering:
            - Timeline modifications
            - Resource reallocation
            - Priority changes
            - Risk mitigation
            - Dependency optimization
            
            Return the updated plan in the same JSON format as the input.
            """
        )
        self.chain = LLMChain(llm=self.llm, prompt=self.prompt)
    
    def optimize_plan(self, current_plan: List[Dict[str, Any]], feedback: str, constraints: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """Optimize project plan based on feedback"""
        try:
            result = self.chain.run(
                current_plan=json.dumps(current_plan, indent=2),
                feedback=feedback,
                constraints=json.dumps(constraints or {}, indent=2)
            )
            
            json_match = re.search(r'\[.*\]', result, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
            else:
                return current_plan
                
        except Exception as e:
            st.error(f"Error optimizing plan: {str(e)}")
            return current_plan

# =============================================================================
# KNOWLEDGE BASE & TEMPLATES
# =============================================================================

class ProjectKnowledgeBase:
    """Simple knowledge base for storing and retrieving project templates"""
    
    def __init__(self, api_key=None):
        # Use simple keyword matching instead of embeddings to avoid async issues
        self.templates = [
            {
                "name": "Web Application Template",
                "type": "web_app",
                "keywords": ["web", "application", "website", "frontend", "backend", "api", "database", "authentication"],
                "tasks": ["Setup", "Backend API", "Frontend UI", "Database", "Authentication", "Testing", "Deployment"],
                "timeline": "4-8 weeks",
                "complexity": "Medium"
            },
            {
                "name": "Mobile App Template",
                "type": "mobile_app",
                "keywords": ["mobile", "app", "android", "ios", "react native", "flutter", "swift", "kotlin"],
                "tasks": ["Setup", "UI Design", "Core Features", "API Integration", "Testing", "App Store Submission"],
                "timeline": "6-12 weeks",
                "complexity": "Medium"
            },
            {
                "name": "API Service Template",
                "type": "api",
                "keywords": ["api", "service", "backend", "rest", "graphql", "microservice", "server"],
                "tasks": ["Architecture", "Database Design", "API Development", "Authentication", "Documentation", "Testing"],
                "timeline": "3-6 weeks",
                "complexity": "Simple"
            },
            {
                "name": "E-commerce Platform Template",
                "type": "ecommerce",
                "keywords": ["ecommerce", "shop", "store", "cart", "payment", "product", "inventory"],
                "tasks": ["Setup", "Product Catalog", "Shopping Cart", "Payment Integration", "User Management", "Order Management", "Testing", "Deployment"],
                "timeline": "8-12 weeks",
                "complexity": "Complex"
            },
            {
                "name": "Data Analysis Dashboard Template",
                "type": "dashboard",
                "keywords": ["dashboard", "analytics", "data", "visualization", "charts", "reporting", "bi"],
                "tasks": ["Setup", "Data Pipeline", "Visualization Components", "Dashboard UI", "Data Processing", "Testing", "Deployment"],
                "timeline": "6-10 weeks",
                "complexity": "Medium"
            }
        ]
    
    def find_similar_templates(self, project_description: str, k: int = 3) -> List[Dict[str, Any]]:
        """Find similar project templates using keyword matching"""
        if not project_description:
            return []
        
        description_lower = project_description.lower()
        template_scores = []
        
        for template in self.templates:
            score = 0
            # Count keyword matches
            for keyword in template["keywords"]:
                if keyword in description_lower:
                    score += 1
            
            # Bonus for exact type matches
            if template["type"] in description_lower:
                score += 2
            
            if score > 0:
                template_scores.append((template, score))
        
        # Sort by score and return top k
        template_scores.sort(key=lambda x: x[1], reverse=True)
        return [template for template, score in template_scores[:k]]

# =============================================================================
# DOCUMENT PROCESSING
# =============================================================================

class DocumentProcessor:
    """Process uploaded documents (PDF, DOCX, TXT)"""
    
    @staticmethod
    def extract_text_from_pdf(pdf_file) -> str:
        """Extract text from PDF file"""
        try:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            st.error(f"Error reading PDF: {str(e)}")
            return ""
    
    @staticmethod
    def extract_text_from_docx(docx_file) -> str:
        """Extract text from DOCX file"""
        try:
            doc = DocxDocument(docx_file)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            st.error(f"Error reading DOCX: {str(e)}")
            return ""
    
    @staticmethod
    def extract_text_from_txt(txt_file) -> str:
        """Extract text from TXT file"""
        try:
            return str(txt_file.read(), "utf-8")
        except Exception as e:
            st.error(f"Error reading TXT: {str(e)}")
            return ""

# =============================================================================
# EXCEL GENERATOR
# =============================================================================

class ExcelGenerator:
    """Generate professional Excel project plans"""
    
    @staticmethod
    def create_excel_plan(tasks: List[Dict[str, Any]], project_metadata: Dict[str, Any]) -> io.BytesIO:
        """Create Excel file with project plan"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Plan"
        
        # Headers
        headers = [
            "Task ID", "Task Name", "Description", "Duration (Days)", 
            "Dependencies", "Start Date", "End Date", "Assigned To", 
            "Priority", "Category", "Status"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add task data
        for row, task in enumerate(tasks, 2):
            ws.cell(row=row, column=1, value=task['id'])
            ws.cell(row=row, column=2, value=task['task_name'])
            ws.cell(row=row, column=3, value=task['description'])
            ws.cell(row=row, column=4, value=task['duration'])
            ws.cell(row=row, column=5, value=', '.join(map(str, task.get('dependencies', []))))
            ws.cell(row=row, column=6, value=task['start_date'])
            ws.cell(row=row, column=7, value=task['end_date'])
            ws.cell(row=row, column=8, value=task['assigned_to'])
            ws.cell(row=row, column=9, value=task['priority'])
            ws.cell(row=row, column=10, value=task['category'])
            ws.cell(row=row, column=11, value="Not Started")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add project metadata sheet
        meta_ws = wb.create_sheet("Project Overview")
        meta_data = [
            ["Project Name", project_metadata.get('name', 'N/A')],
            ["Project Type", project_metadata.get('type', 'N/A')],
            ["Total Duration", project_metadata.get('duration', 'N/A')],
            ["Complexity", project_metadata.get('complexity', 'N/A')],
            ["Team Size", project_metadata.get('team_size', 'N/A')],
            ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row, (key, value) in enumerate(meta_data, 1):
            meta_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            meta_ws.cell(row=row, column=2, value=value)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

# =============================================================================
# MAIN APPLICATION
# =============================================================================

class AgenticProjectPlannerApp:
    """Main Streamlit application"""
    
    def __init__(self):
        self.setup_page_config()
        self.initialize_session_state()
        self.setup_agents()
    
    def setup_page_config(self):
        """Configure Streamlit page"""
        st.set_page_config(
            page_title="Agentic AI Project Planner",
            page_icon="🧠",
            layout="wide",
            initial_sidebar_state="expanded"
        )
    
    def initialize_session_state(self):
        """Initialize session state variables"""
        if 'project_data' not in st.session_state:
            st.session_state.project_data = None
        if 'tasks' not in st.session_state:
            st.session_state.tasks = []
        if 'project_metadata' not in st.session_state:
            st.session_state.project_metadata = {}
    
    def setup_agents(self):
        """Initialize LLM and agents"""
        # Get Gemini API key
        gemini_api_key = st.sidebar.text_input("Gemini API Key", type="password")
        
        if gemini_api_key:
            try:
                # Configure Gemini API
                genai.configure(api_key=gemini_api_key)
                
                # Initialize LLM with proper parameters
                llm = ChatGoogleGenerativeAI(
                    model="gemini-2.0-flash",
                    google_api_key=gemini_api_key,
                    temperature=0.7,
                    convert_system_message_to_human=True
                )
                
                self.parser_agent = ProjectParserAgent(llm)
                self.planner_agent = ProjectPlannerAgent(llm)
                self.optimizer_agent = ProjectOptimizerAgent(llm)
                
                # Initialize simple knowledge base (no embeddings to avoid async issues)
                self.knowledge_base = ProjectKnowledgeBase(gemini_api_key)
                
                self.doc_processor = DocumentProcessor()
                st.sidebar.success("✅ Agents initialized successfully!")
            except Exception as e:
                st.sidebar.error(f"❌ Error initializing agents: {str(e)}")
                self.parser_agent = None
        else:
            st.sidebar.warning("Please enter your Gemini API Key to use the AI agents")
            self.parser_agent = None
    
    def render_header(self):
        """Render application header"""
        st.title("🧠 Agentic AI Project Planner")
        st.markdown("### Intelligent project planning with autonomous agents")
        st.markdown("---")
    
    def render_input_section(self):
        """Render project input section"""
        st.header("📝 Project Requirements Input")
        
        input_method = st.radio(
            "Choose input method:",
            ["Natural Language", "Structured Form", "Upload Document"],
            horizontal=True
        )
        
        if input_method == "Natural Language":
            requirements = st.text_area(
                "Describe your project:",
                placeholder="Example: Build a web application with user authentication, a dashboard for managing tasks, analytics for tracking progress, and a mobile-responsive design. The project should be completed in 4-6 weeks with a team of 3-4 developers.",
                height=150
            )
            
            col1, col2, col3 = st.columns([1, 1, 2])
            with col1:
                team_size = st.number_input("Team Size", min_value=1, max_value=20, value=3)
            with col2:
                priority = st.selectbox("Priority", ["High", "Medium", "Low"])
            
            if st.button("🚀 Generate Project Plan", type="primary"):
                if requirements and self.parser_agent:
                    self.process_requirements(requirements, team_size)
                elif not self.parser_agent:
                    st.error("Please configure your Gemini API key first")
                else:
                    st.error("Please enter project requirements")
        
        elif input_method == "Structured Form":
            self.render_structured_form()
        
        else:  # Upload Document
            self.render_document_upload()
    
    def render_structured_form(self):
        """Render structured input form"""
        with st.form("structured_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                project_name = st.text_input("Project Name")
                project_type = st.selectbox(
                    "Project Type",
                    ["Web Application", "Mobile App", "Desktop Software", "API/Backend", "Other"]
                )
                timeline = st.selectbox(
                    "Expected Timeline",
                    ["2-4 weeks", "1-2 months", "3-6 months", "6+ months"]
                )
            
            with col2:
                team_size = st.number_input("Team Size", min_value=1, max_value=20, value=3)
                complexity = st.selectbox("Complexity", ["Simple", "Medium", "Complex"])
                priority = st.selectbox("Priority", ["High", "Medium", "Low"])
            
            features = st.text_area("Key Features", placeholder="List the main features...")
            constraints = st.text_area("Constraints", placeholder="Budget, timeline, technical constraints...")
            
            if st.form_submit_button("Generate Plan", type="primary"):
                if project_name and features and self.parser_agent:
                    # Convert form data to natural language
                    requirements = f"""
                    Project Name: {project_name}
                    Type: {project_type}
                    Timeline: {timeline}
                    Team Size: {team_size}
                    Complexity: {complexity}
                    Priority: {priority}
                    Features: {features}
                    Constraints: {constraints}
                    """
                    self.process_requirements(requirements, team_size)
                elif not self.parser_agent:
                    st.error("Please configure your Gemini API key first")
                else:
                    st.error("Please fill in required fields")
    
    def render_document_upload(self):
        """Render document upload section"""
        uploaded_file = st.file_uploader(
            "Upload project document",
            type=['pdf', 'docx', 'txt'],
            help="Upload a project brief, requirements document, or specification"
        )
        
        if uploaded_file:
            file_type = uploaded_file.type
            
            if file_type == "application/pdf":
                text = self.doc_processor.extract_text_from_pdf(uploaded_file)
            elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                text = self.doc_processor.extract_text_from_docx(uploaded_file)
            elif file_type == "text/plain":
                text = self.doc_processor.extract_text_from_txt(uploaded_file)
            else:
                st.error("Unsupported file type")
                return
            
            if text:
                st.text_area("Extracted Text", text, height=200, disabled=True)
                
                col1, col2 = st.columns([1, 3])
                with col1:
                    team_size = st.number_input("Team Size", min_value=1, max_value=20, value=3)
                
                if st.button("Generate Plan from Document", type="primary"):
                    if self.parser_agent:
                        self.process_requirements(text, team_size)
                    else:
                        st.error("Please configure your Gemini API key first")
    
    def process_requirements(self, requirements: str, team_size: int):
        """Process requirements and generate project plan"""
        with st.spinner("🤖 AI agents are analyzing your project..."):
            # Step 1: Parse requirements
            project_data = self.parser_agent.parse_requirements(requirements)
            st.session_state.project_data = project_data
            
            # Step 2: Create project plan
            tasks = self.planner_agent.create_project_plan(project_data, team_size)
            st.session_state.tasks = tasks
            
            # Step 3: Create metadata
            total_duration = max([
                (datetime.strptime(task['end_date'], '%Y-%m-%d') - 
                 datetime.strptime(task['start_date'], '%Y-%m-%d')).days + 1
                for task in tasks
            ]) if tasks else 0
            
            st.session_state.project_metadata = {
                'name': project_data.get('project_name', 'Custom Project'),
                'type': project_data.get('project_type', 'Unknown'),
                'duration': f"{total_duration} days",
                'complexity': project_data.get('complexity', 'Medium'),
                'team_size': team_size
            }
        
        st.success("✅ Project plan generated successfully!")
        st.rerun()
    
    def render_project_overview(self):
        """Render project overview section"""
        if not st.session_state.project_metadata:
            return
        
        st.header("📊 Project Overview")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Project Name", st.session_state.project_metadata.get('name', 'N/A'))
        with col2:
            st.metric("Duration", st.session_state.project_metadata.get('duration', 'N/A'))
        with col3:
            st.metric("Complexity", st.session_state.project_metadata.get('complexity', 'N/A'))
        with col4:
            st.metric("Team Size", st.session_state.project_metadata.get('team_size', 'N/A'))
    
    def render_task_list(self):
        """Render generated tasks"""
        if not st.session_state.tasks:
            return
        
        st.header("📋 Generated Tasks")
        
        # Task summary
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Tasks", len(st.session_state.tasks))
        with col2:
            high_priority = len([t for t in st.session_state.tasks if t.get('priority') == 'High'])
            st.metric("High Priority", high_priority)
        with col3:
            total_days = sum([t.get('duration', 0) for t in st.session_state.tasks])
            st.metric("Total Effort", f"{total_days} days")
        
        # Tasks table
        df = pd.DataFrame(st.session_state.tasks)
        
        # Reorder columns for better display
        column_order = ['id', 'task_name', 'description', 'duration', 'start_date', 'end_date', 'assigned_to', 'priority', 'category']
        df = df.reindex(columns=[col for col in column_order if col in df.columns])
        
        st.dataframe(
            df,
            use_container_width=True,
            column_config={
                "id": "ID",
                "task_name": "Task Name",
                "description": "Description",
                "duration": "Duration (Days)",
                "start_date": "Start Date",
                "end_date": "End Date",
                "assigned_to": "Assigned To",
                "priority": "Priority",
                "category": "Category"
            }
        )
        
        # Export options
        col1, col2, col3 = st.columns([1, 1, 2])
        
        with col1:
            if st.button("📥 Export to Excel", type="primary"):
                excel_file = ExcelGenerator.create_excel_plan(
                    st.session_state.tasks, 
                    st.session_state.project_metadata
                )
                
                st.download_button(
                    label="Download Excel File",
                    data=excel_file.getvalue(),
                    file_name=f"project_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            if st.button("📊 Generate Gantt Chart"):
                st.info("Gantt chart generation would be implemented here")
    
    def render_feedback_section(self):
        """Render feedback and optimization section"""
        if not st.session_state.tasks:
            return
        
        st.header("🔄 Plan Optimization")
        
        feedback = st.text_area(
            "What would you like to adjust?",
            placeholder="Example: The frontend development seems too aggressive. Can we extend it by 2-3 days? Also, we need to add a testing phase before deployment.",
            height=100
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            timeline_adjustment = st.selectbox(
                "Timeline Adjustment",
                ["No change", "Extend by 1 week", "Extend by 2 weeks", "Compress timeline"]
            )
        
        with col2:
            priority_change = st.selectbox(
                "Priority Focus",
                ["No change", "Prioritize backend", "Focus on UI/UX", "Emphasize testing"]
            )
        
        if st.button("🔄 Optimize Plan", type="primary"):
            if feedback and self.optimizer_agent:
                constraints = {
                    "timeline_adjustment": timeline_adjustment,
                    "priority_change": priority_change
                }
                
                with st.spinner("🤖 Optimizing project plan..."):
                    optimized_tasks = self.optimizer_agent.optimize_plan(
                        st.session_state.tasks, feedback, constraints
                    )
                    st.session_state.tasks = optimized_tasks
                    st.success("✅ Plan optimized successfully!")
                    st.rerun()
            elif not feedback:
                st.error("Please provide feedback for optimization")
            else:
                st.error("Please configure your Gemini API key first")
    
    def render_analytics_section(self):
        """Render project analytics and insights"""
        if not st.session_state.tasks:
            return
        
        st.header("📈 Project Analytics")
        
        tasks = st.session_state.tasks
        df = pd.DataFrame(tasks)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Tasks by Category")
            if 'category' in df.columns:
                category_counts = df['category'].value_counts()
                st.bar_chart(category_counts)
            
            st.subheader("Priority Distribution")
            if 'priority' in df.columns:
                priority_counts = df['priority'].value_counts()
                st.bar_chart(priority_counts)
        
        with col2:
            st.subheader("Resource Allocation")
            if 'assigned_to' in df.columns:
                resource_counts = df['assigned_to'].value_counts()
                st.bar_chart(resource_counts)
            
            st.subheader("Duration Analysis")
            if 'duration' in df.columns:
                avg_duration = df['duration'].mean()
                max_duration = df['duration'].max()
                min_duration = df['duration'].min()
                
                col_a, col_b, col_c = st.columns(3)
                with col_a:
                    st.metric("Avg Duration", f"{avg_duration:.1f} days")
                with col_b:
                    st.metric("Max Duration", f"{max_duration} days")
                with col_c:
                    st.metric("Min Duration", f"{min_duration} days")
        
        # Critical Path Analysis
        st.subheader("Critical Path Analysis")
        critical_tasks = [task for task in tasks if task.get('priority') == 'High']
        if critical_tasks:
            st.write(f"Found {len(critical_tasks)} critical tasks:")
            critical_df = pd.DataFrame(critical_tasks)[['task_name', 'duration', 'start_date', 'end_date']]
            st.dataframe(critical_df, use_container_width=True)
        else:
            st.info("No critical tasks identified")
    
    def render_templates_section(self):
        """Render project templates and knowledge base"""
        st.header("📚 Project Templates")
        
        if hasattr(self, 'knowledge_base') and self.knowledge_base:
            search_query = st.text_input("Search for similar project templates:", 
                                       placeholder="e.g., web application with authentication")
            
            if search_query:
                similar_templates = self.knowledge_base.find_similar_templates(search_query)
                
                if similar_templates:
                    st.subheader("Similar Templates Found:")
                    for i, template in enumerate(similar_templates, 1):
                        with st.expander(f"{i}. {template['name']}"):
                            st.write(f"**Type:** {template['type']}")
                            st.write(f"**Timeline:** {template['timeline']}")
                            st.write(f"**Complexity:** {template['complexity']}")
                            st.write(f"**Tasks:** {', '.join(template['tasks'])}")
                            
                            if st.button(f"Use Template {i}", key=f"use_template_{i}"):
                                st.info("Template application would be implemented here")
                else:
                    st.info("No similar templates found. Try different keywords like 'web app', 'mobile', 'api', 'dashboard', or 'ecommerce'")
            else:
                # Show all available templates when no search query
                st.subheader("Available Templates:")
                for i, template in enumerate(self.knowledge_base.templates, 1):
                    with st.expander(f"{i}. {template['name']}"):
                        st.write(f"**Type:** {template['type']}")
                        st.write(f"**Keywords:** {', '.join(template['keywords'])}")
                        st.write(f"**Timeline:** {template['timeline']}")
                        st.write(f"**Complexity:** {template['complexity']}")
                        st.write(f"**Tasks:** {', '.join(template['tasks'])}")
        else:
            st.warning("Knowledge base not available. Please configure Gemini API key.")
    
    def render_collaboration_section(self):
        """Render team collaboration features"""
        if not st.session_state.tasks:
            return
        
        st.header("👥 Team Collaboration")
        
        # Team member view
        tasks = st.session_state.tasks
        team_members = list(set([task.get('assigned_to', 'Unassigned') for task in tasks]))
        
        selected_member = st.selectbox("View tasks for team member:", team_members)
        
        if selected_member:
            member_tasks = [task for task in tasks if task.get('assigned_to') == selected_member]
            
            if member_tasks:
                st.subheader(f"Tasks for {selected_member}")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Tasks", len(member_tasks))
                with col2:
                    total_effort = sum([task.get('duration', 0) for task in member_tasks])
                    st.metric("Total Effort", f"{total_effort} days")
                with col3:
                    high_priority = len([task for task in member_tasks if task.get('priority') == 'High'])
                    st.metric("High Priority", high_priority)
                
                # Task list for selected member
                member_df = pd.DataFrame(member_tasks)
                display_columns = ['task_name', 'description', 'duration', 'start_date', 'end_date', 'priority', 'category']
                available_columns = [col for col in display_columns if col in member_df.columns]
                st.dataframe(member_df[available_columns], use_container_width=True)
            else:
                st.info(f"No tasks assigned to {selected_member}")
        
        # Task status updates
        st.subheader("Task Status Updates")
        task_to_update = st.selectbox(
            "Select task to update:",
            options=[f"{task['id']}: {task['task_name']}" for task in tasks],
            format_func=lambda x: x.split(': ', 1)[1] if ': ' in x else x
        )
        
        if task_to_update:
            new_status = st.selectbox(
                "Update status:",
                ["Not Started", "In Progress", "Completed", "Blocked", "On Hold"]
            )
            
            notes = st.text_area("Add notes:", placeholder="Any updates or blockers...")
            
            if st.button("Update Task Status"):
                task_id = int(task_to_update.split(':')[0])
                # In a real application, this would update the task in the database
                st.success(f"Task {task_id} status updated to '{new_status}'")
                if notes:
                    st.info(f"Notes added: {notes}")
    
    def render_risk_assessment(self):
        """Render risk assessment and mitigation suggestions"""
        if not st.session_state.tasks:
            return
        
        st.header("⚠️ Risk Assessment")
        
        tasks = st.session_state.tasks
        
        # Identify potential risks
        risks = []
        
        # Check for tasks with high duration
        high_duration_tasks = [task for task in tasks if task.get('duration', 0) > 7]
        if high_duration_tasks:
            risks.append({
                "type": "High Duration Tasks",
                "severity": "Medium",
                "description": f"{len(high_duration_tasks)} tasks with duration > 7 days",
                "mitigation": "Consider breaking down into smaller tasks"
            })
        
        # Check for resource overallocation
        resource_allocation = {}
        for task in tasks:
            assigned_to = task.get('assigned_to', 'Unassigned')
            if assigned_to not in resource_allocation:
                resource_allocation[assigned_to] = 0
            resource_allocation[assigned_to] += task.get('duration', 0)
        
        overallocated = {k: v for k, v in resource_allocation.items() if v > 40}
        if overallocated:
            risks.append({
                "type": "Resource Overallocation",
                "severity": "High",
                "description": f"Resources with >40 days: {', '.join(overallocated.keys())}",
                "mitigation": "Redistribute tasks or add team members"
            })
        
        # Check for dependency chains
        complex_dependencies = [task for task in tasks if len(task.get('dependencies', [])) > 2]
        if complex_dependencies:
            risks.append({
                "type": "Complex Dependencies",
                "severity": "Medium",
                "description": f"{len(complex_dependencies)} tasks with >2 dependencies",
                "mitigation": "Review dependency structure for optimization"
            })
        
        if risks:
            st.subheader("Identified Risks")
            for i, risk in enumerate(risks, 1):
                severity_color = {"High": "🔴", "Medium": "🟡", "Low": "🟢"}
                with st.expander(f"{severity_color.get(risk['severity'], '⚪')} {risk['type']} ({risk['severity']} Risk)"):
                    st.write(f"**Description:** {risk['description']}")
                    st.write(f"**Mitigation:** {risk['mitigation']}")
        else:
            st.success("✅ No significant risks identified")
        
        # Risk mitigation suggestions
        st.subheader("General Risk Mitigation Strategies")
        mitigation_strategies = [
            "🔄 **Regular Check-ins:** Schedule daily/weekly standups to track progress",
            "📋 **Clear Documentation:** Maintain updated requirements and specifications",
            "🧪 **Testing Strategy:** Implement continuous testing throughout development",
            "💬 **Communication Plan:** Establish clear communication channels",
            "📊 **Progress Tracking:** Use project management tools for visibility",
            "🎯 **Scope Management:** Regularly review and control scope creep"
        ]
        
        for strategy in mitigation_strategies:
            st.markdown(strategy)
    
    def run(self):
        """Main application runner"""
        self.render_header()
        
        # Sidebar navigation
        st.sidebar.title("Navigation")
        page = st.sidebar.radio(
            "Go to:",
            [
                "📝 Project Input",
                "📊 Overview",
                "📋 Tasks",
                "🔄 Optimization",
                "📈 Analytics",
                "📚 Templates",
                "👥 Collaboration",
                "⚠️ Risk Assessment"
            ]
        )
        
        # Render selected page
        if page == "📝 Project Input":
            self.render_input_section()
        elif page == "📊 Overview":
            self.render_project_overview()
        elif page == "📋 Tasks":
            self.render_task_list()
        elif page == "🔄 Optimization":
            self.render_feedback_section()
        elif page == "📈 Analytics":
            self.render_analytics_section()
        elif page == "📚 Templates":
            self.render_templates_section()
        elif page == "👥 Collaboration":
            self.render_collaboration_section()
        elif page == "⚠️ Risk Assessment":
            self.render_risk_assessment()
        
        # Footer
        st.markdown("---")
        st.markdown("### 🚀 Powered by Agentic AI")
        st.markdown("Intelligent project planning with autonomous agents for optimized workflows")

# =============================================================================
# GANTT CHART GENERATOR (Additional Feature)
# =============================================================================

class GanttChartGenerator:
    """Generate Gantt charts for project visualization"""
    
    @staticmethod
    def create_gantt_chart(tasks: List[Dict[str, Any]]) -> None:
        """Create and display a Gantt chart using Plotly"""
        try:
            import plotly.express as px
            import plotly.graph_objects as go
            from datetime import datetime
            
            # Prepare data for Gantt chart
            gantt_data = []
            for task in tasks:
                start_date = datetime.strptime(task['start_date'], '%Y-%m-%d')
                end_date = datetime.strptime(task['end_date'], '%Y-%m-%d')
                
                gantt_data.append({
                    'Task': task['task_name'],
                    'Start': start_date,
                    'Finish': end_date,
                    'Resource': task.get('assigned_to', 'Unassigned'),
                    'Category': task.get('category', 'Other'),
                    'Priority': task.get('priority', 'Medium')
                })
            
            # Create Gantt chart
            df_gantt = pd.DataFrame(gantt_data)
            
            # Color mapping for categories
            color_map = {
                'Setup': '#FF6B6B',
                'Design': '#4ECDC4',
                'Backend': '#45B7D1',
                'Frontend': '#96CEB4',
                'Testing': '#FFEAA7',
                'Deployment': '#DDA0DD',
                'Other': '#95A5A6'
            }
            
            fig = px.timeline(
                df_gantt,
                x_start="Start",
                x_end="Finish",
                y="Task",
                color="Category",
                color_discrete_map=color_map,
                title="Project Gantt Chart",
                hover_data=["Resource", "Priority"]
            )
            
            fig.update_layout(
                height=max(400, len(tasks) * 40),
                xaxis_title="Timeline",
                yaxis_title="Tasks",
                showlegend=True
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
        except ImportError:
            st.error("Plotly not installed. Please install plotly to view Gantt charts.")
        except Exception as e:
            st.error(f"Error creating Gantt chart: {str(e)}")

# =============================================================================
# ADVANCED FEATURES
# =============================================================================

class ProjectEstimator:
    """Advanced project estimation using historical data and machine learning"""
    
    def __init__(self, llm):
        self.llm = llm
        self.prompt = PromptTemplate(
            input_variables=["project_features", "similar_projects"],
            template="""
            Based on the project features and similar historical projects, provide an estimation:
            
            Project Features: {project_features}
            Similar Projects: {similar_projects}
            
            Estimate:
            1. Development time (in person-days)
            2. Potential risks and uncertainties
            3. Recommended team composition
            4. Budget estimation (if possible)
            5. Success probability
            
            Return in JSON format:
            {{
                "estimated_days": 120,
                "confidence_level": "Medium",
                "risks": [...],
                "team_recommendation": [...],
                "budget_range": "...",
                "success_probability": "75%"
            }}
            """
        )
        self.chain = LLMChain(llm=self.llm, prompt=self.prompt)
    
    def estimate_project(self, project_data: Dict[str, Any], similar_projects: List[Dict] = None) -> Dict[str, Any]:
        """Provide advanced project estimation"""
        try:
            result = self.chain.run(
                project_features=json.dumps(project_data, indent=2),
                similar_projects=json.dumps(similar_projects or [], indent=2)
            )
            
            json_match = re.search(r'\{.*\}', result, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
            else:
                return self._fallback_estimation()
                
        except Exception as e:
            st.error(f"Error in project estimation: {str(e)}")
            return self._fallback_estimation()
    
    def _fallback_estimation(self) -> Dict[str, Any]:
        """Fallback estimation when AI fails"""
        return {
            "estimated_days": 90,
            "confidence_level": "Medium",
            "risks": ["Scope creep", "Technical complexity"],
            "team_recommendation": ["Full-stack developer", "Designer", "QA engineer"],
            "budget_range": "$50,000 - $80,000",
            "success_probability": "70%"
        }

# =============================================================================
# APPLICATION ENTRY POINT
# =============================================================================

def main():
    """Main application entry point"""
    app = AgenticProjectPlannerApp()
    app.run()

if __name__ == "__main__":
    main()

# =============================================================================
# USAGE INSTRUCTIONS
# =============================================================================

"""
SETUP INSTRUCTIONS:

.......
"""